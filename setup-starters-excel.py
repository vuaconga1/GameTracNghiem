# -*- coding: utf-8 -*-
"""Tạo 6 sheet Starters riêng + HuongDan trong Game Trắc Nghiệm.xlsx"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

XLSX = Path(__file__).resolve().parent / "Game Trắc Nghiệm.xlsx"
COURSE = "Global Success Lớp 1 - Học kỳ 2"

WORD_BANK_NV = "fifteen|brother|sister|grandmother|shirts|shoes|shorts|tent|teapot|blanket"
WORD_BANK_DH = "fifteen|brother|blanket|tent"
WORD_BANK_KT = "yogurt|yams|yo-yos|zoo|zebu|zebra|sliding|riding|driving|grapes"

# (tab_name, headers, note_row, sample_rows, col_widths, extra_validations)
# extra_validations: list of (col_letter, formula1)
SHEETS = [
    (
        "Nhin_va_viet",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Tiêu đề bài", "Hướng dẫn",
            "Hộp từ gợi ý", "Link hình ảnh", "Đáp án đúng", "Đang dùng",
        ],
        [
            "HK2-NV-01", "1, 2, 3…", "Trùng sheet Khóa học", "Dòng 1 mã bài",
            "Dòng 1 mã bài", "từ1|từ2|từ3", "Link Google Drive", "Từ đúng", "Có/Không",
        ],
        [
            [
                "HK2-NV-01", 1, COURSE,
                "STARTERS ENGLISH TEST - TERM 2", "Look and write.",
                WORD_BANK_NV, "https://drive.google.com/file/d/MAU-15/view", "fifteen", "Có",
            ],
            [
                "HK2-NV-01", 2, COURSE, "", "", "",
                "https://drive.google.com/file/d/MAU-brother/view", "brother", "Có",
            ],
            [
                "HK2-NV-01", 3, COURSE, "", "", "",
                "https://drive.google.com/file/d/MAU-sister/view", "sister", "Có",
            ],
        ],
        [12, 8, 28, 24, 18, 36, 32, 14, 12],
        [],
    ),
    (
        "Chon_va_khoanh",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Link hình ảnh",
            "Lựa chọn A", "Lựa chọn B", "Đáp án đúng", "Đang dùng",
        ],
        [
            "HK2-CK-01", "1, 2, 3…", "Trùng sheet Khóa học", "Link Google Drive",
            "Từ 1", "Từ 2", "Từ đúng hoặc A/B", "Có/Không",
        ],
        [
            [
                "HK2-CK-01", 1, COURSE,
                "https://drive.google.com/file/d/MAU-boy/view",
                "brother", "sister", "brother", "Có",
            ],
            [
                "HK2-CK-01", 2, COURSE,
                "https://drive.google.com/file/d/MAU-tent/view",
                "teapot", "tent", "tent", "Có",
            ],
        ],
        [12, 8, 28, 32, 14, 14, 14, 12],
        [],
    ),
    (
        "Doc_va_hoan_thanh",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Hộp từ gợi ý", "Nội dung câu",
            "Link hình gợi ý", "Đáp án đúng", "Đang dùng",
        ],
        [
            "HK2-DH-01", "1, 2, 3…", "Trùng sheet Khóa học", "Dòng 1: từ1|từ2",
            "Câu tiếng Anh, ___ = chỗ trống", "Ảnh gợi ý (tuỳ chọn)", "Từ đúng", "Có/Không",
        ],
        [
            [
                "HK2-DH-01", 1, COURSE, WORD_BANK_DH,
                "I am ___ years old.", "https://drive.google.com/file/d/MAU-num15/view",
                "fifteen", "Có",
            ],
            [
                "HK2-DH-01", 2, COURSE, "",
                "I sleep under a warm ___.", "https://drive.google.com/file/d/MAU-blanket/view",
                "blanket", "Có",
            ],
        ],
        [12, 8, 28, 32, 28, 28, 14, 12],
        [],
    ),
    (
        "Doc_va_noi",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Nội dung câu", "Link hình ảnh",
            "Nhãn hình", "Đáp án đúng", "Đang dùng",
        ],
        [
            "HK2-DN-01", "1, 2, 3…", "Trùng sheet Khóa học", "Câu bên trái",
            "Hình bên phải", "A, B, C…", "Nhãn đúng", "Có/Không",
        ],
        [
            [
                "HK2-DN-01", 1, COURSE, "I wear my new ___.",
                "https://drive.google.com/file/d/MAU-shoes/view", "A", "A", "Có",
            ],
            [
                "HK2-DN-01", 2, COURSE, "My grandma is very kind.",
                "https://drive.google.com/file/d/MAU-grandma/view", "B", "B", "Có",
            ],
            [
                "HK2-DN-01", 3, COURSE, "I like camping in a ___.",
                "https://drive.google.com/file/d/MAU-tent2/view", "C", "C", "Có",
            ],
        ],
        [12, 8, 28, 28, 32, 10, 12, 12],
        [],
    ),
    (
        "Kiem_tra_tu_vung",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Hướng dẫn", "Hộp từ gợi ý",
            "Link hình ảnh", "Đáp án đúng", "Đang dùng",
        ],
        [
            "HK2-KT-01", "1, 2, 3…", "Trùng sheet Khóa học", "Dòng 1 mã bài",
            "Dòng 1: từ1|từ2", "Link Google Drive", "Từ đúng", "Có/Không",
        ],
        [
            [
                "HK2-KT-01", 1, COURSE,
                "Look at the pictures and write the correct words. Use the words in the box.",
                WORD_BANK_KT, "https://drive.google.com/file/d/MAU-yogurt/view", "yogurt", "Có",
            ],
            [
                "HK2-KT-01", 2, COURSE, "", "",
                "https://drive.google.com/file/d/MAU-yams/view", "yams", "Có",
            ],
        ],
        [12, 8, 28, 36, 36, 32, 14, 12],
        [],
    ),
    (
        "Kiem_tra_dung_sai",
        [
            "Mã bài", "Thứ tự", "Tên khóa học", "Link hình ảnh", "Từ hiển thị",
            "Câu mô tả", "Đúng hay sai?", "Đang dùng",
        ],
        [
            "HK2-DS-01", "1, 2, 3…", "Trùng sheet Khóa học", "Link Google Drive",
            "Từ tiếng Anh", "Câu tiếng Anh", "Đúng hoặc Sai", "Có/Không",
        ],
        [
            [
                "HK2-DS-01", 1, COURSE,
                "https://drive.google.com/file/d/MAU-question/view",
                "question", "This is a question.", "Đúng", "Có",
            ],
            [
                "HK2-DS-01", 2, COURSE,
                "https://drive.google.com/file/d/MAU-ox/view",
                "fox", "This is a fox.", "Sai", "Có",
            ],
            [
                "HK2-DS-01", 3, COURSE,
                "https://drive.google.com/file/d/MAU-fox/view",
                "ox", "This is an ox.", "Sai", "Có",
            ],
        ],
        [12, 8, 28, 32, 14, 24, 14, 12],
        [("G", '"Đúng,Sai"')],
    ),
]

HUONG_DAN_LINES = [
    "HUONG DAN NHAP CAU HOI STARTERS",
    "",
    "Moi LOAI GAME co mot tab rieng - mo dung tab can nhap.",
    "Moi bai choi = nhieu dong cung Ma bai, Thu tu = 1, 2, 3...",
    "Hop tu: cach nhau bang dau |  (VD: cat|dog|bird)",
    "Link anh: Drive -> Chia se -> Bat ky ai co lien ket.",
    "",
    "CAC TAB:",
    "  Nhin_va_viet       -> Nhin va viet",
    "  Chon_va_khoanh     -> Chon va khoanh",
    "  Doc_va_hoan_thanh  -> Doc va hoan thanh",
    "  Doc_va_noi         -> Doc va noi",
    "  Kiem_tra_tu_vung   -> Kiem tra tu vung",
    "  Kiem_tra_dung_sai  -> Kiem tra dung sai",
    "",
    "Chi tiet: DATABASE_STARTERS.md",
]

OLD_SHEETS_TO_REMOVE = ["CauHoi_Starters"]


def remove_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


def setup_data_sheet(wb, spec):
    name, headers, notes, samples, widths, extra_vals = spec
    remove_sheet_if_exists(wb, name)
    ws = wb.create_sheet(name)

    header_fill = PatternFill("solid", fgColor="E8EEF8")
    note_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(bold=True, color="0D2B6E")
    note_font = Font(italic=True, color="666666", size=10)

    ncol = len(headers)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, note in enumerate(notes, start=1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.font = note_font
        cell.fill = note_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r_idx, data in enumerate(samples, start=3):
        for c_idx, val in enumerate(data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = 500
    dang_dung_col = get_column_letter(ncol)
    dv_active = DataValidation(type="list", formula1='"Có,Không"', allow_blank=True)
    ws.add_data_validation(dv_active)
    dv_active.add(f"{dang_dung_col}3:{dang_dung_col}{last_row}")

    for col_letter, formula in extra_vals:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}3:{col_letter}{last_row}")


def setup_huong_dan(wb):
    remove_sheet_if_exists(wb, "HuongDan")
    ws = wb.create_sheet("HuongDan")
    title_font = Font(bold=True, size=14, color="0D2B6E")
    body_font = Font(size=11)
    for i, line in enumerate(HUONG_DAN_LINES, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = title_font if i == 1 else body_font
    ws.column_dimensions["A"].width = 72


def main():
    if not XLSX.exists():
        raise SystemExit("xlsx not found")

    backup = XLSX.with_suffix(".backup-starters-v2.xlsx")
    shutil.copy2(XLSX, backup)

    wb = openpyxl.load_workbook(XLSX)
    for old in OLD_SHEETS_TO_REMOVE:
        remove_sheet_if_exists(wb, old)

    setup_huong_dan(wb)
    total_rows = 0
    for spec in SHEETS:
        setup_data_sheet(wb, spec)
        total_rows += len(spec[3])

    out = XLSX.parent / "Game_Trac_Nghiem_starters_v2.xlsx"
    wb.save(out)


if __name__ == "__main__":
    main()
